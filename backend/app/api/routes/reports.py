# backend/app/api/routes/reports.py
# Paste this ENTIRE file into: backend/app/api/routes/reports.py
# Requires: pip install openpyxl

import os
from typing import Optional
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.core.database import get_db
from app.core.security import get_current_user

router = APIRouter()


# ── Dashboard stats (used by Reports page KPI cards) ─────────────────────────
@router.get('/summary')
def report_summary(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Return contract count summary grouped by status"""
    result = db.execute(text("""
        SELECT
            COUNT(*)                                                        AS total_contracts,
            SUM(CASE WHEN status = 'draft'           THEN 1 ELSE 0 END)    AS drafts,
            SUM(CASE WHEN status = 'in_review'       THEN 1 ELSE 0 END)    AS in_review,
            SUM(CASE WHEN status = 'vendor_feedback' THEN 1 ELSE 0 END)    AS vendor_feedback,
            SUM(CASE WHEN status = 'approved'        THEN 1 ELSE 0 END)    AS approved,
            SUM(CASE WHEN status = 'signed'          THEN 1 ELSE 0 END)    AS signed
        FROM contracts
    """))
    row = result.mappings().first()
    return dict(row) if row else {}


# ── Export to Excel ───────────────────────────────────────────────────────────
@router.get('/export')
def export_report(
    status:    Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date:   Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Export filtered contracts to an Excel (.xlsx) file.
    Query params:
        status    — filter by contract status
        from_date — YYYY-MM-DD start date filter
        to_date   — YYYY-MM-DD end date filter
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=500,
            detail='openpyxl not installed. Run: pip install openpyxl'
        )

    # ── Build query ───────────────────────────────────────────────────────────
    query  = "SELECT * FROM contracts WHERE 1=1"
    params = {}

    if status:
        query += " AND status = :status"
        params['status'] = status

    if from_date:
        query += " AND DATE(created_at) >= :from_date"
        params['from_date'] = from_date

    if to_date:
        query += " AND DATE(created_at) <= :to_date"
        params['to_date'] = to_date

    query += " ORDER BY created_at DESC"

    result    = db.execute(text(query), params)
    contracts = result.mappings().all()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contract Report'

    # Styles
    header_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill    = PatternFill('solid', fgColor='1A2980')          # ContractPro navy
    subheader_fill = PatternFill('solid', fgColor='26D0CE')          # ContractPro cyan
    center_align   = Alignment(horizontal='center', vertical='center')
    wrap_align     = Alignment(wrap_text=True, vertical='top')
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    title_cell       = ws['A1']
    title_cell.value = 'ContractPro — Contract Activity Report'
    title_cell.font  = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
    title_cell.fill  = PatternFill('solid', fgColor='1A2980')
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 32

    # ── Generated-on row ─────────────────────────────────────────────────────
    ws.merge_cells('A2:J2')
    meta_cell       = ws['A2']
    meta_cell.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   "
        f"Status filter: {status or 'All'}   |   "
        f"Date range: {from_date or 'All'} → {to_date or 'All'}   |   "
        f"Total records: {len(contracts)}"
    )
    meta_cell.font      = Font(name='Calibri', italic=True, size=10, color='FFFFFF')
    meta_cell.fill      = PatternFill('solid', fgColor='26D0CE')
    meta_cell.alignment = center_align
    ws.row_dimensions[2].height = 20

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        'Contract #', 'Title', 'Category', 'Status',
        'Vendor', 'Value (BDT)', 'SLA Days', 'Handler',
        'Uploaded By', 'Created Date',
    ]
    col_widths = [14, 35, 18, 14, 25, 15, 10, 22, 20, 14]

    for col_idx, header in enumerate(headers, start=1):
        cell              = ws.cell(row=3, column=col_idx, value=header)
        cell.font         = header_font
        cell.fill         = PatternFill('solid', fgColor='2C3E50')
        cell.alignment    = center_align
        cell.border       = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    ws.row_dimensions[3].height = 20

    # Status badge colors for Excel cells
    STATUS_COLORS = {
        'draft':           'ADB5BD',
        'in_review':       'FFC107',
        'vendor_feedback': '0DCAF0',
        'approved':        '198754',
        'signed':          '6F42C1',
    }

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, contract in enumerate(contracts, start=4):
        c          = dict(contract)
        row_fill   = PatternFill('solid', fgColor='F8F9FA') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        status_val = c.get('status', '')
        status_clr = STATUS_COLORS.get(status_val, 'EEEEEE')

        row_data = [
            c.get('contract_number', f"CTR-{c.get('id', '')}"),
            c.get('title', ''),
            (c.get('category') or '').replace('_', ' ').title(),
            status_val.replace('_', ' ').title(),
            c.get('vendor_name') or c.get('vendor') or '',
            float(c['contract_value']) if c.get('contract_value') else '',
            c.get('sla_days', 7),
            c.get('current_handler', ''),
            c.get('uploaded_by', ''),
            str(c['created_at'])[:10] if c.get('created_at') else '',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = wrap_align
            # Highlight status column
            if col_idx == 4:
                cell.fill = PatternFill('solid', fgColor=status_clr)
                cell.font = Font(bold=True, color='FFFFFF' if status_val in ('approved','signed') else '000000')
            else:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 18

    # ── Freeze panes at row 4 ─────────────────────────────────────────────────
    ws.freeze_panes = 'A4'

    # ── Save to buffer and stream ─────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"contractpro_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )